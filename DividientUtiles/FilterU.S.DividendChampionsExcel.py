from dataclasses import dataclass
from typing import Optional, Tuple, Dict, Callable, List
import operator
import pandas as pd
from openpyxl import load_workbook
from openpyxl.utils import get_column_letter
from openpyxl.styles import Alignment, Font, PatternFill
from enum import Enum
import math

# === Enums ===
class StyleAttr(Enum):
    FONT_COLOR = "font_color"
    BOLD = "bold"
    FILL_COLOR = "fill_color"

class Colors(Enum):
    RED_COLOR = "EA3323"
    BLUE_COLOR = "2F6EBA"
    GRAY_COLOR = "808080"

# === Data model for column settings ===
@dataclass
class ColumnSettings:
    column_name: str
    filter: Optional[Tuple[Callable, float]] = None
    style: Optional[Dict[str, object]] = None
    conditional_style: Optional[Dict[str, object]] = None

# === Configuration ===
MARKET_DIVIDEND_YIELD = 1.34
UPPER_LIMIT_FOR_DIVIDEND_YIELD = 2
FREEZING_CELL = "C2"

column_definitions: List[ColumnSettings] = [
    ColumnSettings("Company Name"),
    ColumnSettings("Ticker Symbol"),
    ColumnSettings("Sector"),
    ColumnSettings("Industry"),
    ColumnSettings("No. Yrs", filter=(operator.gt, 25), style={StyleAttr.BOLD.value: True}),
    ColumnSettings("Price", style={StyleAttr.FONT_COLOR.value: Colors.BLUE_COLOR.value}),
    ColumnSettings("Div. Yield", filter=(operator.gt, MARKET_DIVIDEND_YIELD),
                    conditional_style={
                        "condition": (operator.gt, UPPER_LIMIT_FOR_DIVIDEND_YIELD),
                        StyleAttr.FONT_COLOR.value: Colors.RED_COLOR.value
                    }),
    ColumnSettings("MR% Inc.", filter=(operator.gt, 6), style={StyleAttr.FONT_COLOR.value: Colors.GRAY_COLOR.value}),
    ColumnSettings("DGR 1-yr", filter=(operator.gt, 6), style={StyleAttr.FONT_COLOR.value: Colors.GRAY_COLOR.value}),
    ColumnSettings("DGR 3-yr", filter=(operator.gt, 6), style={StyleAttr.FONT_COLOR.value: Colors.GRAY_COLOR.value}),
    ColumnSettings("DGR 5-yr", filter=(operator.gt, 6), style={StyleAttr.FONT_COLOR.value: Colors.GRAY_COLOR.value}),
    ColumnSettings("DGR 10-yr", filter=(operator.gt, 6), style={StyleAttr.FONT_COLOR.value: Colors.GRAY_COLOR.value}),
    ColumnSettings("EPS% Payout", filter=(operator.le, 60)),
    ColumnSettings("Past 5yr Growth", filter=(operator.gt, 0), style={StyleAttr.FONT_COLOR.value: Colors.BLUE_COLOR.value}),
    ColumnSettings("Est-5yr Growth", filter=(operator.gt, 0), style={StyleAttr.FONT_COLOR.value: Colors.BLUE_COLOR.value}),
    ColumnSettings("MktCap ($Mil)", style={StyleAttr.FONT_COLOR.value: Colors.BLUE_COLOR.value}),
    ColumnSettings("Debt/ Equity", style={StyleAttr.FONT_COLOR.value: Colors.BLUE_COLOR.value}),
]

def main():
    file_path = input("Enter Excel file name (without extension): ").strip() + ".xlsx"
    sheet_name = input("Enter sheet name (default is 'All CCC'): ").strip() or "All CCC"

    data_df, combined_columns = read_excel_data(file_path, sheet_name)
    matched = match_columns(data_df)
    filtered_df = apply_filters(data_df, matched)
    filtered_df = round_numeric_columns(filtered_df)
    new_file = save_filtered_df(filtered_df, file_path)

    format_excel_file(new_file, matched)

    # Validation steps
    valid_filter = validate_filtered_rows(filtered_df, matched, column_definitions)
    if valid_filter:
        print("✅ Filtering is reliable and works as expected.")
    else:
        print("❌ There are issues with filtering. Please check the messages above.")

    print(f"\u2705 Done! Saved filtered and formatted data to: {new_file}")

def read_excel_data(file_path, sheet_name):
    raw_header = pd.read_excel(file_path, sheet_name=sheet_name, header=[4, 5], nrows=0)
    combined_columns = [' '.join(map(str, col)).strip() for col in raw_header.columns.values]
    data_df = pd.read_excel(file_path, sheet_name=sheet_name, header=None, skiprows=6)
    data_df.columns = combined_columns
    return data_df, combined_columns

def match_columns(data_df):
    import difflib
    matched = {}
    for col_def in column_definitions:
        matches = [c for c in data_df.columns if col_def.column_name.lower() in c.lower()]
        if matches:
            matched[col_def.column_name] = matches[0]
        else:
            suggestion = difflib.get_close_matches(col_def.column_name, data_df.columns, n=1)
            print(f"Warning: Column '{col_def.column_name}' not found. Suggested: {suggestion}")
    return matched

def apply_filters(data_df, matched):
    filtered_df = data_df.copy()
    for col_def in column_definitions:
        if col_def.filter and col_def.column_name in matched:
            col = matched[col_def.column_name]
            op_func, val = col_def.filter
            try:
                filtered_df = filtered_df[filtered_df[col].apply(
                    lambda x: op_func(float(str(x).replace('%', '').replace(',', '')), val)
                    if pd.notnull(x) else False)]
            except Exception as e:
                print(f"Error filtering column '{col}': {e}")
    final_cols = [matched[c.column_name] for c in column_definitions if c.column_name in matched]
    return filtered_df[final_cols]

def round_numeric_columns(df):
    for col in df.select_dtypes(include=['float', 'int']).columns:
        df[col] = df[col].map(lambda x: round(x, 2) if isinstance(x, float) else x)
    return df

def save_filtered_df(filtered_df, original_file):
    import os
    base, ext = os.path.splitext(original_file)
    new_file = f"{base}-Filtered.xlsx"
    filtered_df.to_excel(new_file, sheet_name="FilteredResult", index=False)
    return new_file

def format_excel_file(file_path, matched):
    wb = load_workbook(file_path)
    ws = wb["FilteredResult"]
    apply_styles(ws, matched)
    auto_adjust_column_widths(ws)
    freeze_header_and_columns(ws)
    apply_thousand_separator(ws)
    wb.save(file_path)

def apply_styles(ws, matched):
    for col_idx, col_def in enumerate(column_definitions, 1):
        if col_def.column_name not in matched:
            continue
        for row in ws.iter_rows(min_row=2, min_col=col_idx, max_col=col_idx):
            for cell in row:
                if col_def.style:
                    apply_cell_style(cell, col_def.style)
                if col_def.conditional_style:
                    condition, threshold = col_def.conditional_style['condition']
                    try:
                        val = float(str(cell.value).replace('%', '').replace(',', ''))
                        if condition(val, threshold):
                            apply_cell_style(cell, col_def.conditional_style)
                    except:
                        continue

def apply_cell_style(cell, style_dict):
    cell.alignment = Alignment(horizontal='center', vertical='center')
    font_color = style_dict.get(StyleAttr.FONT_COLOR.value, None)
    bold = style_dict.get(StyleAttr.BOLD.value, False)
    fill_color = style_dict.get(StyleAttr.FILL_COLOR.value, None)

    if bold and font_color:
        cell.font = Font(bold=True, color=font_color)
    elif bold:
        cell.font = Font(bold=True)
    elif font_color:
        cell.font = Font(color=font_color)

    if fill_color:
        cell.fill = PatternFill(start_color=fill_color, end_color=fill_color, fill_type="solid")

def auto_adjust_column_widths(ws):
    for col_cells in ws.columns:
        max_len = max((len(str(cell.value)) for cell in col_cells if cell.value is not None), default=0)
        ws.column_dimensions[get_column_letter(col_cells[0].column)].width = max_len + 2

def freeze_header_and_columns(ws):
    ws.freeze_panes = FREEZING_CELL

def apply_thousand_separator(ws):
    for row in ws.iter_rows(min_row=2):
        for cell in row:
            if cell.value is None:
                continue
            if isinstance(cell.value, (int, float)):
                try:
                    if abs(cell.value) > 999:
                        # If integer, no decimals shown
                        if isinstance(cell.value, int) or (isinstance(cell.value, float) and cell.value.is_integer()):
                            cell.number_format = '#,##0'
                        else:
                            # Decimal number with 2 digits after decimal point
                            cell.number_format = '#,##0.00'
                except Exception:
                    pass

# === Validation ===

def validate_filtered_rows(df_filtered, matched_columns, column_definitions, tolerance=1e-3):
    all_valid = True
    changed_values = {}
    columns_with_rounded_differences = set()

    for col_def in column_definitions:
        if col_def.filter and col_def.column_name in matched_columns:
            col_name = matched_columns[col_def.column_name]
            op_func, threshold = col_def.filter

            for idx, val in df_filtered[col_name].items():
                try:
                    raw_val = str(val).replace('%', '').replace(',', '')
                    numeric_val = float(raw_val)

                    if not op_func(numeric_val, threshold):
                        print(f"❌ Row {idx} failed filter for column '{col_def.column_name}': value = {val}")
                        all_valid = False

                    original_val = numeric_val
                    if not math.isclose(float(val), original_val, abs_tol=tolerance):
                        changed_values.setdefault(col_name, []).append((idx, original_val))
                        columns_with_rounded_differences.add(col_name)

                except Exception as e:
                    print(f"⚠️ Skipping validation for row {idx} in column '{col_def.column_name}': {e}")

    if all_valid:
        print("✅ All filtered rows meet filter conditions.")
    else:
        print("❌ There are issues with filtering. Please check the messages above.")

    for col, diffs in changed_values.items():
        print(f"⚠️ Values in column '{col}' changed after filtering (likely due to rounding):")
        for idx, val in diffs:
            print(f"   Row {idx}: {val}")

    if columns_with_rounded_differences:
        print("📝 Note: The following columns have values that were slightly changed (rounded):")
        for col in columns_with_rounded_differences:
            print(f"   - {col}")

    return all_valid



if __name__ == "__main__":
    main()
